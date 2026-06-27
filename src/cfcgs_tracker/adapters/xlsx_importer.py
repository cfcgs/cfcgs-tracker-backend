import re
import unicodedata
from io import BytesIO
from typing import Any, Iterator

from openpyxl import load_workbook


def normalize_header(header: Any) -> str:
    text = "" if header is None else str(header).strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = text.lower()
    text = text.replace("&", " and ")
    text = re.sub(r"[%/()-]", " ", text)
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def iter_sheet_row_batches(
    file_bytes: bytes,
    *,
    sheet_index: int,
    batch_size: int,
) -> Iterator[list[dict[str, Any]]]:
    workbook = load_workbook(
        BytesIO(file_bytes),
        data_only=True,
        read_only=True,
    )
    worksheet = workbook.worksheets[sheet_index]
    rows = worksheet.iter_rows(values_only=True)

    header_row = next(rows, None)
    if header_row is None:
        workbook.close()
        return

    headers = [normalize_header(value) for value in header_row]
    batch: list[dict[str, Any]] = []

    try:
        for row in rows:
            if row is None or not any(
                value not in (None, "") for value in row
            ):
                continue

            batch.append(
                {
                    header: value
                    for header, value in zip(headers, row, strict=False)
                    if header
                }
            )

            if len(batch) >= batch_size:
                yield batch
                batch = []

        if batch:
            yield batch
    finally:
        workbook.close()
