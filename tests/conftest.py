from contextlib import contextmanager
from datetime import datetime
from io import BytesIO, StringIO
import csv

import factory
import pytest
import pytest_asyncio
from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy import event
from sqlalchemy.ext.asyncio import AsyncSession, create_async_engine
from testcontainers.postgres import PostgresContainer

from src.cfcgs_tracker.adapters.orm import get_session
from src.cfcgs_tracker.domain.base import table_registry
from src.cfcgs_tracker.domain.models import (
    BeneficiaryCountry,
    ClimateFinanceRecord,
    FundFocus,
    FundType,
    FundingProvider,
    Project,
    ProviderFundProfile,
    Source,
    User,
    UserRole,
)
from src.cfcgs_tracker.service_layer.security import get_password_hash
from src.cfcgs_tracker.settings import Settings
from src.cfcgs_tracker.app import app


@pytest.fixture
def client(session):
    async def get_session_override():
        yield session

    with TestClient(app) as client:
        app.dependency_overrides[get_session] = get_session_override
        yield client

    app.dependency_overrides.clear()


@pytest.fixture(scope="session")
def engine():
    with PostgresContainer("postgres:17", driver="psycopg") as postgres:
        yield create_async_engine(postgres.get_connection_url())


@pytest_asyncio.fixture
async def session(engine):
    async with engine.begin() as conn:
        await conn.run_sync(table_registry.metadata.create_all)

    async with AsyncSession(engine, expire_on_commit=False) as session:
        yield session

    async with engine.begin() as conn:
        await conn.run_sync(table_registry.metadata.drop_all)


@contextmanager
def _mock_db_time(
    *,
    model=User,
    created_time=datetime(2026, 5, 31),
    updated_time=datetime(2026, 6, 1),
):
    def create_fake_time_hook(mapper, connection, target):
        if hasattr(target, "created_at"):
            target.created_at = created_time
            target.updated_at = created_time

    def update_fake_time_hook(mapper, connection, target):
        if hasattr(target, "updated_at"):
            target.updated_at = updated_time

    event.listen(model, "before_insert", create_fake_time_hook)
    event.listen(model, "before_update", update_fake_time_hook)

    yield created_time, updated_time

    event.remove(model, "before_insert", create_fake_time_hook)
    event.remove(model, "before_update", update_fake_time_hook)


@pytest.fixture
def mock_db_time():
    return _mock_db_time


@pytest_asyncio.fixture
async def user(session: AsyncSession):
    password = "testtest"
    user = UserFactory(
        password=get_password_hash(password),
        role=UserRole.importer,
    )
    session.add(user)
    await session.commit()
    await session.refresh(user)

    user.clean_password = password

    return user


@pytest_asyncio.fixture
async def other_user(session: AsyncSession):
    password = "testtest"
    user = UserFactory(
        password=get_password_hash(password),
        role=UserRole.importer,
    )
    session.add(user)
    await session.commit()
    await session.refresh(user)

    user.clean_password = password

    return user


@pytest.fixture
def token(client, user):
    response = client.post(
        "/auth/token",
        data={
            "username": user.email,
            "password": user.clean_password,
        },
    )

    return response.json()["access_token"]


@pytest.fixture
def other_token(client, other_user):
    response = client.post(
        "/auth/token",
        data={
            "username": other_user.email,
            "password": other_user.clean_password,
        },
    )

    return response.json()["access_token"]


@pytest_asyncio.fixture
async def admin_user(session: AsyncSession):
    password = "testtest"
    user = UserFactory(
        password=get_password_hash(password),
        role=UserRole.admin,
    )
    session.add(user)
    await session.commit()
    await session.refresh(user)

    user.clean_password = password

    return user


@pytest.fixture
def admin_token(client, admin_user):
    response = client.post(
        "/auth/token",
        data={
            "username": admin_user.email,
            "password": admin_user.clean_password,
        },
    )

    return response.json()["access_token"]


@pytest.fixture
def settings():
    settings = Settings()
    return settings


class UserFactory(factory.Factory):
    class Meta:
        model = User

    username = factory.Sequence(lambda n: f"test-user{n}")
    email = factory.LazyAttribute(lambda obj: f"{obj.username}@test.com")

    password = factory.LazyAttribute(
        lambda obj: f"{obj.username}+asujdahZia123"
    )
    role = UserRole.importer


class FundTypeFactory(factory.Factory):
    class Meta:
        model = FundType

    name = factory.Sequence(lambda n: f"fund-type-{n}")


class FundFocusFactory(factory.Factory):
    class Meta:
        model = FundFocus

    name = factory.Sequence(lambda n: f"fund-focus-{n}")


class BeneficiaryCountryFactory(factory.Factory):
    class Meta:
        model = BeneficiaryCountry

    name = factory.Sequence(lambda n: f"country-{n}")


class FundingProviderFactory(factory.Factory):
    class Meta:
        model = FundingProvider

    name = factory.Sequence(lambda n: f"provider-{n}")


class ProviderFundProfileFactory(factory.Factory):
    class Meta:
        model = ProviderFundProfile

    funding_provider_id = 1
    fund_type_id = None
    fund_focus_id = None
    pledge = 100.0
    deposit = 80.0
    approval = 60.0
    disbursement = 40.0
    projects_approved = 5


class ProjectFactory(factory.Factory):
    class Meta:
        model = Project

    title = factory.Sequence(lambda n: f"project-{n}")


class SourceFactory(factory.Factory):
    class Meta:
        model = Source

    name = factory.Sequence(lambda n: f"source-{n}")
    url = factory.LazyAttribute(lambda obj: f"https://{obj.name}.org")


class ClimateFinanceRecordFactory(factory.Factory):
    class Meta:
        model = ClimateFinanceRecord

    year = 2022
    source_id = 1
    source_row_hash = factory.Sequence(lambda n: f"hash-{n}")
    approved_amount_usd_millions = 10.0
    climate_finance_amount_usd_millions = 8.0
    adaptation_amount_usd_millions = 3.0
    mitigation_amount_usd_millions = 2.0
    both_objectives_amount_usd_millions = 3.0
    project_id = None
    funding_provider_id = None
    financial_instrument_id = None
    sector_id = None
    sub_sector_id = None
    beneficiary_country_id = None


@pytest_asyncio.fixture
async def fund_type_factory(session: AsyncSession):
    async def create_fund_type(**kwargs) -> FundType:
        fund_type = FundTypeFactory(**kwargs)
        session.add(fund_type)
        await session.flush()
        return fund_type

    return create_fund_type


@pytest_asyncio.fixture
async def fund_focus_factory(session: AsyncSession):
    async def create_fund_focus(**kwargs) -> FundFocus:
        fund_focus = FundFocusFactory(**kwargs)
        session.add(fund_focus)
        await session.flush()
        return fund_focus

    return create_fund_focus


@pytest_asyncio.fixture
async def beneficiary_country_factory(session: AsyncSession):
    async def create_beneficiary_country(**kwargs) -> BeneficiaryCountry:
        country = BeneficiaryCountryFactory(**kwargs)
        session.add(country)
        await session.flush()
        return country

    return create_beneficiary_country


@pytest_asyncio.fixture
async def funding_provider_factory(session: AsyncSession):
    async def create_funding_provider(**kwargs) -> FundingProvider:
        provider = FundingProviderFactory(**kwargs)
        session.add(provider)
        await session.flush()
        return provider

    return create_funding_provider


@pytest_asyncio.fixture
async def project_factory(session: AsyncSession):
    async def create_project(**kwargs) -> Project:
        project = ProjectFactory(**kwargs)
        session.add(project)
        await session.flush()
        return project

    return create_project


@pytest_asyncio.fixture
async def source_factory(session: AsyncSession):
    async def create_source(**kwargs) -> Source:
        source = SourceFactory(**kwargs)
        session.add(source)
        await session.flush()
        return source

    return create_source


@pytest_asyncio.fixture
async def climate_finance_record_factory(session: AsyncSession):
    async def create_climate_finance_record(**kwargs) -> ClimateFinanceRecord:
        record = ClimateFinanceRecordFactory(**kwargs)
        session.add(record)
        await session.flush()
        return record

    return create_climate_finance_record


@pytest_asyncio.fixture
async def provider_fund_profile_factory(session: AsyncSession):
    async def create_provider_fund_profile(
        *,
        funding_provider: FundingProvider,
        fund_type: FundType | None = None,
        fund_focus: FundFocus | None = None,
        **kwargs,
    ) -> ProviderFundProfile:
        profile = ProviderFundProfileFactory(
            funding_provider_id=funding_provider.id,
            fund_type_id=fund_type.id if fund_type else None,
            fund_focus_id=fund_focus.id if fund_focus else None,
            **kwargs,
        )
        session.add(profile)
        await session.flush()
        return profile

    return create_provider_fund_profile


class ClimateFinanceImportRowFactory(factory.DictFactory):
    year = factory.Faker("random_int", min=2020, max=2026)
    project_title = factory.Faker("word")
    beneficiary_country = factory.Faker("word")
    provider_type = factory.Iterator(["Bilateral", "Multilateral", "Other"])
    funding_provider = factory.Faker("word")
    approved_amount_usd_millions = factory.Faker(
        "pyfloat", positive=True, min_value=10, max_value=500, right_digits=2
    )
    climate_finance_amount_usd_millions = factory.Faker(
        "pyfloat", positive=True, min_value=5, max_value=400, right_digits=2
    )
    adaptation_climate_finance_amount_usd_millions = factory.Faker(
        "pyfloat", positive=True, min_value=1, max_value=100, right_digits=2
    )
    mitigation_climate_finance_amount_usd_millions = factory.Faker(
        "pyfloat", positive=True, min_value=1, max_value=100, right_digits=2
    )
    both_climate_objectives_finance_amount_usd_millions = factory.Faker(
        "pyfloat", positive=True, min_value=1, max_value=100, right_digits=2
    )
    sector = factory.Faker("word")
    sub_sector = factory.Faker("word")
    financial_instrument = factory.Iterator(["Grant", "Loan", "Equity"])
    source = factory.Iterator(["OECD", "CFU", "IDB"])
    source_url = factory.Faker("uri")


class FundImportRowFactory(factory.DictFactory):
    fund = factory.Faker("word")
    fund_type = factory.Iterator(["Multilateral", "Bilateral", "Regional"])
    fund_focus = factory.Iterator(["Adaptation", "Mitigation", "Crosscutting"])
    pledge_usd_mn = factory.Faker(
        "pyfloat", positive=True, min_value=10, max_value=800, right_digits=2
    )
    deposit_usd_mn = factory.Faker(
        "pyfloat", positive=True, min_value=10, max_value=700, right_digits=2
    )
    approval_usd_mn = factory.Faker(
        "pyfloat", positive=True, min_value=10, max_value=600, right_digits=2
    )
    disbursement_usd_mn = factory.Faker(
        "pyfloat", positive=True, min_value=10, max_value=500, right_digits=2
    )
    number_of_projects_approved = factory.Faker("random_int", min=1, max=50)


def build_xlsx_file(
    *,
    climate_rows: list[dict] | None = None,
    fund_rows: list[dict] | None = None,
    file_name: str | None = "test_import.xlsx",
) -> tuple[str, bytes, str]:
    climate_rows = climate_rows or [ClimateFinanceImportRowFactory()]
    fund_rows = fund_rows or [FundImportRowFactory()]

    workbook = Workbook()

    sheet1 = workbook.active
    sheet1.title = "sheet1"
    sheet1.append(["dummy"])

    sheet2 = workbook.create_sheet("sheet2")
    sheet2.append(["dummy"])

    sheet3 = workbook.create_sheet("sheet3")
    sheet3.append(["dummy"])

    climate_sheet = workbook.create_sheet("climate_finance")
    climate_headers = [
        "Year",
        "Project Title",
        "Beneficiary Country",
        "Provider Type",
        "Funding Provider",
        "Approved Amount (USD millions)",
        "Climate Finance Amount (USD millions)",
        "Adaptation Climate Finance Amount (USD millions)",
        "Mitigation Climate Finance Amount (USD millions)",
        "Both Climate Objectives Finance Amount (USD millions)",
        "Sector",
        "Sub-sector",
        "Financial Instrument",
        "Source",
        "Source URL",
    ]
    climate_sheet.append(climate_headers)
    for row in climate_rows:
        climate_sheet.append(
            [
                row["year"],
                row["project_title"],
                row["beneficiary_country"],
                row["provider_type"],
                row["funding_provider"],
                row["approved_amount_usd_millions"],
                row["climate_finance_amount_usd_millions"],
                row["adaptation_climate_finance_amount_usd_millions"],
                row["mitigation_climate_finance_amount_usd_millions"],
                row["both_climate_objectives_finance_amount_usd_millions"],
                row["sector"],
                row["sub_sector"],
                row["financial_instrument"],
                row["source"],
                row["source_url"],
            ]
        )

    fund_sheet = workbook.create_sheet("fund_status")
    fund_headers = [
        "Fund",
        "Fund Type",
        "Fund focus",
        "Pledge (USD mn)",
        "Deposit (USD mn)",
        "Approval (USD mn)",
        "Disbursement (USD mn)",
        "Number of projects approved",
    ]
    fund_sheet.append(fund_headers)
    for row in fund_rows:
        fund_sheet.append(
            [
                row["fund"],
                row["fund_type"],
                row["fund_focus"],
                row["pledge_usd_mn"],
                row["deposit_usd_mn"],
                row["approval_usd_mn"],
                row["disbursement_usd_mn"],
                row["number_of_projects_approved"],
            ]
        )

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    return (
        file_name,
        buffer.read(),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@pytest.fixture
def file_xlsx():
    return build_xlsx_file


def build_csv_file(
    *,
    file_kind: str,
    climate_rows: list[dict] | None = None,
    fund_rows: list[dict] | None = None,
    file_name: str | None = None,
) -> tuple[str, bytes, str]:
    climate_rows = climate_rows or [ClimateFinanceImportRowFactory()]
    fund_rows = fund_rows or [FundImportRowFactory()]

    buffer = StringIO(newline="")

    if file_kind == "climate_finance":
        headers = [
            "Year",
            "Project Title",
            "Beneficiary Country",
            "Provider Type",
            "Funding Provider",
            "Approved Amount (USD millions)",
            "Climate Finance Amount (USD millions)",
            "Adaptation Climate Finance Amount (USD millions)",
            "Mitigation Climate Finance Amount (USD millions)",
            "Both Climate Objectives Finance Amount (USD millions)",
            "Sector",
            "Sub-sector",
            "Financial Instrument",
            "Source",
            "Source URL",
        ]
        writer = csv.writer(buffer)
        writer.writerow(headers)
        for row in climate_rows:
            writer.writerow(
                [
                    row["year"],
                    row["project_title"],
                    row["beneficiary_country"],
                    row["provider_type"],
                    row["funding_provider"],
                    row["approved_amount_usd_millions"],
                    row["climate_finance_amount_usd_millions"],
                    row["adaptation_climate_finance_amount_usd_millions"],
                    row["mitigation_climate_finance_amount_usd_millions"],
                    row["both_climate_objectives_finance_amount_usd_millions"],
                    row["sector"],
                    row["sub_sector"],
                    row["financial_instrument"],
                    row["source"],
                    row["source_url"],
                ]
            )
        return (
            file_name or "test_climate_finance.csv",
            buffer.getvalue().encode("utf-8"),
            "text/csv",
        )

    if file_kind == "fund":
        headers = [
            "Fund",
            "Fund Type",
            "Fund focus",
            "Pledge (USD mn)",
            "Deposit (USD mn)",
            "Approval (USD mn)",
            "Disbursement (USD mn)",
            "Number of projects approved",
        ]
        writer = csv.writer(buffer)
        writer.writerow(headers)
        for row in fund_rows:
            writer.writerow(
                [
                    row["fund"],
                    row["fund_type"],
                    row["fund_focus"],
                    row["pledge_usd_mn"],
                    row["deposit_usd_mn"],
                    row["approval_usd_mn"],
                    row["disbursement_usd_mn"],
                    row["number_of_projects_approved"],
                ]
            )
        return (
            file_name or "test_fund_status.csv",
            buffer.getvalue().encode("utf-8"),
            "text/csv",
        )

    raise ValueError("file_kind must be 'climate_finance' or 'fund'")


@pytest.fixture
def file_csv():
    return build_csv_file
